from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Iterable
import re

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


ABAS_FIXAS = {
    "Dashboard",
    "02_Cronograma",
    "Banco_Ferramentas",
    "Mapa_Selecao",
    "Indice",
    "Relatorio_Consolidado",
    "G1_Relatorio",
    "G2_Relatorio",
    "G3_Relatorio",
    "G4_Relatorio",
    "G5_Relatorio",
    "G1_GoNoGo",
    "G2_GoNoGo",
    "G3_GoNoGo",
    "G4_GoNoGo",
    "G5_GoNoGo",
    "README_GERADOR",
}


def sanitizar_nome_arquivo(texto: str) -> str:
    texto = str(texto or "").strip()
    texto = re.sub(r"[^\w\s\-]", "", texto, flags=re.UNICODE)
    texto = re.sub(r"\s+", "_", texto)
    return texto[:80] or "produto"


def normalizar(texto: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(texto or "").lower())


def is_go_no_go_tool_id(tool_id: str, tool_name: str = "") -> bool:
    texto = f"{tool_id} {tool_name}".lower()
    return any(
        t in texto
        for t in [
            "gonogo",
            "go/no-go",
            "go / no-go",
            "go no-go",
            "no-go",
            "no go",
            "matriz go",
        ]
    )


def to_excel_date(value: date | datetime) -> datetime:
    if isinstance(value, datetime):
        return value
    return datetime(value.year, value.month, value.day)


def format_date(value: date | datetime) -> str:
    return to_excel_date(value).strftime("%d/%m/%Y")


def diff_days(start: date | datetime, end: date | datetime) -> int:
    start_dt = to_excel_date(start)
    end_dt = to_excel_date(end)
    return max((end_dt - start_dt).days + 1, 1)


def add_days(base: date | datetime, days: int) -> datetime:
    return to_excel_date(base) + timedelta(days=days)


def set_value_next_to_label(ws, labels: list[str], value, max_rows: int = 40, max_cols: int = 12) -> bool:
    """
    Procura um rótulo na planilha e preenche a célula à direita.
    Isso deixa o código menos dependente de endereço fixo.
    """
    labels_norm = [normalizar(label) for label in labels]

    for row in range(1, min(ws.max_row, max_rows) + 1):
        for col in range(1, min(ws.max_column, max_cols) + 1):
            cell_value = ws.cell(row=row, column=col).value
            if not cell_value:
                continue

            cell_norm = normalizar(cell_value)
            if any(label in cell_norm or cell_norm in label for label in labels_norm):
                ws.cell(row=row, column=col + 1).value = value
                return True

    return False


def preencher_fallback(ws, cell: str, value) -> None:
    ws[cell] = value


def preencher_campo(ws, labels: list[str], fallback_cell: str, value) -> None:
    ok = set_value_next_to_label(ws, labels, value)
    if not ok:
        preencher_fallback(ws, fallback_cell, value)


def encontrar_aba_por_codigo(wb, codigo: str) -> str | None:
    if codigo in wb.sheetnames:
        return codigo

    codigo_norm = normalizar(codigo)

    for aba in wb.sheetnames:
        if codigo_norm and codigo_norm in normalizar(aba):
            return aba

    return None


def ler_banco_ferramentas(wb) -> list[dict]:
    if "Banco_Ferramentas" not in wb.sheetnames:
        return []

    ws = wb["Banco_Ferramentas"]

    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[normalizar(value)] = col

    col_codigo = headers.get("codigo") or headers.get("cod") or headers.get("id")
    col_ferramenta = headers.get("ferramenta") or headers.get("nome") or headers.get("tool")
    col_gate = headers.get("gate")
    col_aba = (
        headers.get("aba")
        or headers.get("nomedaaba")
        or headers.get("abamodelo")
        or headers.get("sheet")
    )

    if not col_codigo:
        return []

    registros = []

    for row in range(2, ws.max_row + 1):
        codigo = ws.cell(row=row, column=col_codigo).value
        if not codigo:
            continue

        registros.append(
            {
                "row": row,
                "codigo": str(codigo),
                "ferramenta": str(ws.cell(row=row, column=col_ferramenta).value)
                if col_ferramenta else "",
                "gate": str(ws.cell(row=row, column=col_gate).value)
                if col_gate else "",
                "aba": str(ws.cell(row=row, column=col_aba).value)
                if col_aba and ws.cell(row=row, column=col_aba).value else "",
            }
        )

    return registros


def tool_gate_number(engine_gate_label: str) -> int:
    match = re.search(r"gate\s*(\d)", str(engine_gate_label).lower())
    if match:
        return int(match.group(1))
    return 5


def preencher_mapa_selecao(wb, codigos_selecionados: set[str]) -> str:
    if "Mapa_Selecao" not in wb.sheetnames:
        return ""

    ws = wb["Mapa_Selecao"]

    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[normalizar(value)] = col

    col_codigo = headers.get("codigo") or headers.get("cod") or headers.get("id") or 1
    col_selecionar = (
        headers.get("selecionar")
        or headers.get("selecionada")
        or headers.get("selecionado")
        or headers.get("bit")
    )

    if not col_selecionar:
        col_selecionar = ws.max_column + 1
        ws.cell(row=1, column=col_selecionar).value = "Selecionar"

    bits = []

    for row in range(2, ws.max_row + 1):
        codigo = ws.cell(row=row, column=col_codigo).value
        if not codigo:
            continue

        codigo = str(codigo)
        bit = "1" if codigo in codigos_selecionados else "0"
        ws.cell(row=row, column=col_selecionar).value = int(bit)
        bits.append(bit)

    return "".join(bits)


def preencher_banco_ferramentas(wb, codigos_selecionados: set[str]) -> None:
    if "Banco_Ferramentas" not in wb.sheetnames:
        return

    ws = wb["Banco_Ferramentas"]

    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[normalizar(value)] = col

    col_codigo = headers.get("codigo") or headers.get("cod") or headers.get("id") or 1
    col_status = headers.get("selecionar") or headers.get("selecionada") or headers.get("status")

    if not col_status:
        col_status = ws.max_column + 1
        ws.cell(row=1, column=col_status).value = "Selecionada"

    for row in range(2, ws.max_row + 1):
        codigo = ws.cell(row=row, column=col_codigo).value
        if not codigo:
            continue

        codigo = str(codigo)
        ws.cell(row=row, column=col_status).value = "Sim" if codigo in codigos_selecionados else "Não"


def preencher_dashboard(
    wb,
    nome_cronograma: str,
    nome_projeto: str,
    nome_produto: str,
    rota: str,
    complexidade: str,
    score: int,
    codigo_binario: str,
    total_ferramentas: int,
    data_inicio: date | datetime,
    data_fim: date | datetime,
) -> None:
    if "Dashboard" not in wb.sheetnames:
        return

    ws = wb["Dashboard"]

    preencher_campo(ws, ["Nome do cronograma", "Cronograma"], "B2", nome_cronograma)
    preencher_campo(ws, ["Nome do projeto", "Projeto"], "B3", nome_projeto)
    preencher_campo(ws, ["Nome base do produto", "Produto"], "B4", nome_produto)
    preencher_campo(ws, ["Data de início", "Inicio", "Início"], "B5", format_date(data_inicio))
    preencher_campo(ws, ["Data final", "Data fim", "Fim"], "B6", format_date(data_fim))
    preencher_campo(ws, ["Rota"], "B7", rota)
    preencher_campo(ws, ["Complexidade"], "B8", complexidade)
    preencher_campo(ws, ["Score"], "B9", score)
    preencher_campo(ws, ["Quantidade de ferramentas", "Ferramentas selecionadas"], "B10", total_ferramentas)
    preencher_campo(ws, ["Código binário", "Codigo binario"], "B11", codigo_binario)
    preencher_campo(ws, ["Gerado em", "Data de geração"], "B12", datetime.now().strftime("%d/%m/%Y %H:%M"))


def distribuir_periodos_por_gate(
    data_inicio: date | datetime,
    data_fim: date | datetime,
    tools_by_gate: dict[int, list],
) -> dict[int, tuple[datetime, datetime]]:
    """
    Distribui o prazo total entre os Gates de forma proporcional ao número de ferramentas,
    mantendo todos os Gates em sequência.
    """
    start = to_excel_date(data_inicio)
    end = to_excel_date(data_fim)

    total_days = max((end - start).days + 1, 5)

    # Sempre considera os 5 Gates, mesmo que um deles tenha poucas ferramentas.
    gate_counts = {
        gate: max(len(tools_by_gate.get(gate, [])), 1)
        for gate in range(1, 6)
    }

    # Pequeno peso adicional nos Gates 3 e 4 porque normalmente concentram geração/seleção.
    gate_weight_boost = {
        1: 1.0,
        2: 1.1,
        3: 1.25,
        4: 1.25,
        5: 1.0,
    }

    weights = {
        gate: gate_counts[gate] * gate_weight_boost[gate]
        for gate in range(1, 6)
    }

    total_weight = sum(weights.values())

    raw_durations = {
        gate: max(1, round(total_days * weights[gate] / total_weight))
        for gate in range(1, 6)
    }

    # Ajusta soma para bater exatamente o prazo.
    diff = total_days - sum(raw_durations.values())
    gates_order = [3, 4, 2, 1, 5]

    i = 0
    while diff != 0:
        gate = gates_order[i % len(gates_order)]
        if diff > 0:
            raw_durations[gate] += 1
            diff -= 1
        elif diff < 0 and raw_durations[gate] > 1:
            raw_durations[gate] -= 1
            diff += 1
        i += 1

    periods = {}
    cursor = start

    for gate in range(1, 6):
        duration = raw_durations[gate]
        gate_start = cursor
        gate_end = cursor + timedelta(days=duration - 1)

        if gate == 5:
            gate_end = end

        periods[gate] = (gate_start, gate_end)
        cursor = gate_end + timedelta(days=1)

    return periods


def distribuir_periodos_ferramentas(
    gate_start: datetime,
    gate_end: datetime,
    tools: list,
) -> list[tuple[datetime, datetime]]:
    """
    Distribui as ferramentas dentro do período do Gate.
    """
    if not tools:
        return []

    total_days = max((gate_end - gate_start).days + 1, 1)
    n = len(tools)

    base = max(total_days // n, 1)
    resto = total_days - base * n

    periods = []
    cursor = gate_start

    for idx in range(n):
        dur = base + (1 if idx < resto else 0)
        tool_start = cursor
        tool_end = min(cursor + timedelta(days=dur - 1), gate_end)
        periods.append((tool_start, tool_end))
        cursor = tool_end + timedelta(days=1)

    return periods


def preencher_cronograma(
    wb,
    nome_cronograma: str,
    nome_projeto: str,
    nome_produto: str,
    data_inicio: date | datetime,
    data_fim: date | datetime,
    selected_tools,
    engine,
    codigo_binario: str,
) -> None:
    if "02_Cronograma" not in wb.sheetnames:
        return

    ws = wb["02_Cronograma"]

    preencher_campo(ws, ["Nome do cronograma", "Cronograma"], "B3", nome_cronograma)
    preencher_campo(ws, ["Nome do projeto", "Projeto"], "B4", nome_projeto)
    preencher_campo(ws, ["Nome base do produto", "Produto"], "B5", nome_produto)
    preencher_campo(ws, ["Data de início", "Inicio", "Início"], "B6", format_date(data_inicio))
    preencher_campo(ws, ["Data final", "Data fim", "Fim"], "B7", format_date(data_fim))
    preencher_campo(ws, ["Duração total", "Duracao total"], "B8", diff_days(data_inicio, data_fim))
    preencher_campo(ws, ["Código binário", "Codigo binario"], "B9", codigo_binario)
    preencher_campo(ws, ["Quantidade de ferramentas", "Ferramentas selecionadas"], "B10", len(selected_tools))

    tools_by_gate: dict[int, list] = defaultdict(list)

    for item in selected_tools:
        gate = engine.tool_gate(item.tool.id)
        tools_by_gate[gate].append(item)

    gate_periods = distribuir_periodos_por_gate(data_inicio, data_fim, tools_by_gate)

    # -------------------------------------------------------------------------
    # Resumo por Gate
    # -------------------------------------------------------------------------
    resumo_start = 13
    resumo_headers = [
        "Gate",
        "Início",
        "Fim",
        "Duração (dias)",
        "Ferramentas",
        "Status",
    ]

    for col, header in enumerate(resumo_headers, start=1):
        cell = ws.cell(row=resumo_start, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for gate in range(1, 6):
        row = resumo_start + gate
        gate_start, gate_end = gate_periods[gate]
        ws.cell(row=row, column=1).value = f"Gate {gate}"
        ws.cell(row=row, column=2).value = gate_start
        ws.cell(row=row, column=3).value = gate_end
        ws.cell(row=row, column=4).value = (gate_end - gate_start).days + 1
        ws.cell(row=row, column=5).value = len(tools_by_gate.get(gate, []))
        ws.cell(row=row, column=6).value = "Não iniciado"

        ws.cell(row=row, column=2).number_format = "dd/mm/yyyy"
        ws.cell(row=row, column=3).number_format = "dd/mm/yyyy"

    # -------------------------------------------------------------------------
    # Cronograma detalhado por ferramenta
    # -------------------------------------------------------------------------
    detail_start = 22
    headers = [
        "Nº",
        "Gate",
        "Código",
        "Ferramenta",
        "Início",
        "Fim",
        "Duração (dias)",
        "Status",
        "Responsável",
        "Observações",
    ]

    # Limpa área anterior
    for row in range(detail_start, max(ws.max_row + 1, detail_start + 200)):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).value = None

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=detail_start, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = detail_start + 1
    seq = 1

    for gate in range(1, 6):
        tools = tools_by_gate.get(gate, [])
        gate_start, gate_end = gate_periods[gate]
        tool_periods = distribuir_periodos_ferramentas(gate_start, gate_end, tools)

        for item, (tool_start, tool_end) in zip(tools, tool_periods):
            ws.cell(row=row, column=1).value = seq
            ws.cell(row=row, column=2).value = f"Gate {gate}"
            ws.cell(row=row, column=3).value = item.tool.id
            ws.cell(row=row, column=4).value = item.tool.name
            ws.cell(row=row, column=5).value = tool_start
            ws.cell(row=row, column=6).value = tool_end
            ws.cell(row=row, column=7).value = (tool_end - tool_start).days + 1
            ws.cell(row=row, column=8).value = "Não iniciado"
            ws.cell(row=row, column=9).value = ""
            ws.cell(row=row, column=10).value = ""

            ws.cell(row=row, column=5).number_format = "dd/mm/yyyy"
            ws.cell(row=row, column=6).number_format = "dd/mm/yyyy"

            row += 1
            seq += 1

    # Validação de status
    status_validation = DataValidation(
        type="list",
        formula1='"Não iniciado,Em andamento,Concluído,Atrasado,Pausado"',
        allow_blank=False,
    )
    ws.add_data_validation(status_validation)
    status_validation.add(f"H{detail_start + 1}:H{row + 100}")

    # Formatação básica
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in range(resumo_start, resumo_start + 6):
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).alignment = Alignment(vertical="center")

    for r in range(detail_start, row):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).alignment = Alignment(vertical="center", wrap_text=True)

    widths = {
        "A": 8,
        "B": 14,
        "C": 16,
        "D": 42,
        "E": 14,
        "F": 14,
        "G": 16,
        "H": 18,
        "I": 22,
        "J": 36,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = f"A{detail_start + 1}"


def preencher_indice(wb, abas_mantidas: list[str]) -> None:
    if "Indice" not in wb.sheetnames:
        return

    ws = wb["Indice"]

    for row in range(4, 300):
        for col in range(1, 5):
            ws.cell(row=row, column=col).value = None

    ws["A3"] = "Aba"
    ws["B3"] = "Status"

    row = 4

    for aba in abas_mantidas:
        ws.cell(row=row, column=1).value = aba
        ws.cell(row=row, column=2).value = "Aplicável ao produto"
        row += 1


def preencher_relatorio_consolidado(
    wb,
    nome_cronograma: str,
    nome_projeto: str,
    nome_produto: str,
    rota: str,
    data_inicio: date | datetime,
    data_fim: date | datetime,
    selected_tools,
    engine,
) -> None:
    if "Relatorio_Consolidado" not in wb.sheetnames:
        return

    ws = wb["Relatorio_Consolidado"]

    preencher_campo(ws, ["Nome do cronograma", "Cronograma"], "B2", nome_cronograma)
    preencher_campo(ws, ["Nome do projeto", "Projeto"], "B3", nome_projeto)
    preencher_campo(ws, ["Nome base do produto", "Produto"], "B4", nome_produto)
    preencher_campo(ws, ["Rota"], "B5", rota)
    preencher_campo(ws, ["Data de início", "Início"], "B6", format_date(data_inicio))
    preencher_campo(ws, ["Data final", "Fim"], "B7", format_date(data_fim))
    preencher_campo(ws, ["Gerado em"], "B8", datetime.now().strftime("%d/%m/%Y"))

    start_row = 12

    for row in range(start_row, max(ws.max_row + 1, start_row + 200)):
        for col in range(1, 6):
            ws.cell(row=row, column=col).value = None

    ws.cell(row=start_row, column=1).value = "Gate"
    ws.cell(row=start_row, column=2).value = "Código"
    ws.cell(row=start_row, column=3).value = "Ferramenta"

    for col in range(1, 4):
        cell = ws.cell(row=start_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    row = start_row + 1

    for item in selected_tools:
        ws.cell(row=row, column=1).value = engine.tool_gate_label(item.tool.id)
        ws.cell(row=row, column=2).value = item.tool.id
        ws.cell(row=row, column=3).value = item.tool.name
        row += 1


def gerar_planilha_projeto(
    template_path: str | Path,
    nome_cronograma: str,
    nome_projeto: str,
    nome_produto: str,
    data_inicio: date | datetime,
    data_fim: date | datetime,
    rota: str,
    complexidade: str,
    score: int,
    selected_tools,
    engine,
) -> tuple[bytes, str]:
    template_path = Path(template_path)

    if not template_path.exists():
        raise FileNotFoundError(f"Planilha-mãe não encontrada: {template_path}")

    wb = load_workbook(template_path)

    selected_tools = [
        item for item in selected_tools
        if not is_go_no_go_tool_id(item.tool.id, item.tool.name)
    ]

    codigos_selecionados = {item.tool.id for item in selected_tools}

    banco = ler_banco_ferramentas(wb)
    abas_ferramentas = set()

    for item in selected_tools:
        codigo = item.tool.id

        aba_do_banco = ""

        for registro in banco:
            if registro["codigo"] == codigo:
                aba_do_banco = registro.get("aba") or ""
                break

        if aba_do_banco and aba_do_banco in wb.sheetnames:
            abas_ferramentas.add(aba_do_banco)
            continue

        aba_encontrada = encontrar_aba_por_codigo(wb, codigo)

        if aba_encontrada:
            abas_ferramentas.add(aba_encontrada)

    abas_para_manter = set(ABAS_FIXAS) | abas_ferramentas

    codigo_binario = preencher_mapa_selecao(wb, codigos_selecionados)

    preencher_banco_ferramentas(wb, codigos_selecionados)

    preencher_dashboard(
        wb=wb,
        nome_cronograma=nome_cronograma,
        nome_projeto=nome_projeto,
        nome_produto=nome_produto,
        rota=rota,
        complexidade=complexidade,
        score=score,
        codigo_binario=codigo_binario,
        total_ferramentas=len(selected_tools),
        data_inicio=data_inicio,
        data_fim=data_fim,
    )

    preencher_cronograma(
        wb=wb,
        nome_cronograma=nome_cronograma,
        nome_projeto=nome_projeto,
        nome_produto=nome_produto,
        data_inicio=data_inicio,
        data_fim=data_fim,
        selected_tools=selected_tools,
        engine=engine,
        codigo_binario=codigo_binario,
    )

    preencher_relatorio_consolidado(
        wb=wb,
        nome_cronograma=nome_cronograma,
        nome_projeto=nome_projeto,
        nome_produto=nome_produto,
        rota=rota,
        data_inicio=data_inicio,
        data_fim=data_fim,
        selected_tools=selected_tools,
        engine=engine,
    )

    # Remove abas não aplicáveis.
    for aba in list(wb.sheetnames):
        if aba not in abas_para_manter and len(wb.sheetnames) > 1:
            del wb[aba]

    abas_existentes_para_manter = [
        aba for aba in wb.sheetnames
        if aba in abas_para_manter
    ]

    preencher_indice(wb, abas_existentes_para_manter)

    if "Dashboard" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Dashboard")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nome_arquivo = (
        f"Planilha_Conceito_"
        f"{sanitizar_nome_arquivo(nome_produto)}_"
        f"{sanitizar_nome_arquivo(nome_projeto)}.xlsx"
    )

    return output.getvalue(), nome_arquivo
